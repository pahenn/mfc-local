#!/usr/bin/env -S uv run --quiet
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
Convert the *new-shape* Driven Brands "Fleet KPI Table Report" xlsx into a CSV
the DuckDB reconciliation picks up as table `driven_brands_kpi_daily`.

This is the sibling of driven_kpi_xlsx_to_csv.py, which handles the OLD shape
(single sheet, weekly grain, 17 cols with product breakdown). Use that one for
files like "Fleet KPI Table Report - April 7, 2026.xlsx".

The new shape (first seen in "… - April 24, 2026.xlsx"):
  - TWO sheets: `Local` and `National` — each becomes rows tagged with a
    `segment` column ('local' / 'national').
  - DAILY grain (the `Date` column is a single day, not a week-ending Saturday).
  - 9 columns, no Region / Year-Month / COSC-FZ / product breakdown.
  - New metrics: `Net Sales - CY` and `Tickets - CY`.

- Picks the most recent .xlsx under ../driven-data/ by default.
- Writes to ../source-data/driven_brands_kpi_daily.csv.
- Normalizes headers to camelCase. Excel date serials -> ISO YYYY-MM-DD.
"""
import csv
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

# New-shape header contract. The xlsx must contain exactly these columns
# (order-independent). `segment` is synthesized from the sheet name, not read.
HEADER_MAP = {
    "Date":               "date",
    "Fleet Account Name": "fleetAccountName",
    "Fleet Account #":    "fleetAccountNumber",
    "Store #":            "storeNumber",
    "Cars Per Day - CY":  "carsPerDay",
    "Discounts $ - CY":   "discountsDollars",
    "Gross Sales - CY":   "grossSales",
    "Net Sales - CY":     "netSales",
    "Tickets - CY":       "tickets",
}
DATE_COLS = {"date"}

# Output column order. `segment` leads so the file is self-describing.
OUT_COLUMNS = [
    "segment",
    "date",
    "fleetAccountName",
    "fleetAccountNumber",
    "storeNumber",
    "carsPerDay",
    "discountsDollars",
    "grossSales",
    "netSales",
    "tickets",
]

HERE = Path(__file__).resolve().parent
DRIVEN_DIR = HERE.parent / "driven-data"
OUT = HERE.parent / "source-data" / "driven_brands_kpi_daily.csv"


def pick_input(argv: list[str]) -> Path:
    if len(argv) > 1:
        p = Path(argv[1]).expanduser().resolve()
        if not p.exists():
            sys.exit(f"not found: {p}")
        return p
    candidates = sorted(DRIVEN_DIR.glob("*.xlsx"))
    if not candidates:
        sys.exit(f"no .xlsx in {DRIVEN_DIR}")
    if len(candidates) > 1:
        print(f"multiple xlsx found, using most recent: {candidates[-1].name}", file=sys.stderr)
    return candidates[-1]


def fmt(value, field: str):
    if value is None:
        return ""
    if field in DATE_COLS:
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        # openpyxl should parse date-styled cells to datetime; if it comes
        # through as a raw serial, fall back.
        if isinstance(value, (int, float)):
            base = datetime(1899, 12, 30)
            return (base.fromordinal(base.toordinal() + int(value))).strftime("%Y-%m-%d")
    return value


def resolve_headers(raw_headers, sheet_name: str) -> list[str | None]:
    """Map raw header cells to output keys; hard-fail on unknowns so a silent
    shape change surfaces instead of dropping a column."""
    resolved, missing = [], []
    for h in raw_headers:
        key = HEADER_MAP.get(str(h).strip()) if h is not None else None
        if key is None:
            missing.append(h)
        resolved.append(key)
    if missing:
        sys.exit(f"[{sheet_name}] unmapped headers: {missing}")
    return resolved


def main() -> None:
    src = pick_input(sys.argv)
    print(f"reading  {src}", file=sys.stderr)
    wb = load_workbook(src, read_only=True, data_only=True)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    total = 0
    with OUT.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(OUT_COLUMNS)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            segment = sheet_name.strip().lower()
            rows_iter = ws.iter_rows(values_only=True)
            raw_headers = next(rows_iter, None)
            if not raw_headers:
                print(f"[{sheet_name}] empty sheet, skipping", file=sys.stderr)
                continue
            resolved = resolve_headers(raw_headers, sheet_name)

            count = 0
            for row in rows_iter:
                if all(v is None for v in row):
                    continue
                record = {"segment": segment}
                for i, v in enumerate(row):
                    field = resolved[i]
                    if field is not None:
                        record[field] = fmt(v, field)
                w.writerow([record.get(col, "") for col in OUT_COLUMNS])
                count += 1
            total += count
            print(f"  [{segment}] {count:,} rows", file=sys.stderr)

    print(f"wrote    {OUT}  ({total:,} rows)", file=sys.stderr)


if __name__ == "__main__":
    main()
