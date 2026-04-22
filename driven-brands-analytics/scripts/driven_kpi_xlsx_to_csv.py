#!/usr/bin/env -S uv run --quiet
# /// script
# requires-python = ">=3.11"
# dependencies = ["openpyxl>=3.1"]
# ///
"""
Convert the Driven Brands "Fleet KPI Table Report" xlsx into a CSV the
dashboard loader picks up as table `driven_brands_kpi`.

- Picks the most recent .xlsx under ../driven-data/ by default.
- Writes to ../source-data/driven_brands_kpi.csv.
- Normalizes headers to camelCase. Excel date serials → ISO YYYY-MM-DD.
"""
import csv
import sys
from datetime import datetime, date
from pathlib import Path

from openpyxl import load_workbook

HEADER_MAP = {
    "Year/Month":               "yearMonth",
    "Date":                     "date",
    "Fleet Account Name":       "fleetAccountName",
    "Fleet Account #":          "fleetAccountNumber",
    "Region":                   "region",
    "Store #":                  "storeNumber",
    "Week Ending":              "weekEnding",
    "COSC/FZ":                  "coscFz",
    "Air Filters - CY":         "airFilters",
    "Cabin Filters - CY":       "cabinFilters",
    "Cars Per Day - CY":        "carsPerDay",
    "Coolant - CY":             "coolant",
    "Discounts $ - CY":         "discountsDollars",
    "Gross Sales - CY":         "grossSales",
    "Oil Changes - CY":         "oilChanges",
    "Wipers - CY":              "wipers",
    "Differential Fluid % - CY": "differentialFluidPct",
}
DATE_COLS = {"date", "weekEnding"}

HERE = Path(__file__).resolve().parent
DRIVEN_DIR = HERE.parent / "driven-data"
OUT = HERE.parent / "source-data" / "driven_brands_kpi.csv"


def pick_input(argv: list[str]) -> Path:
    if len(argv) > 1:
        p = Path(argv[1]).expanduser().resolve()
        if not p.exists():
            sys.exit(f"not found: {p}")
        return p
    candidates = sorted(DRIVEN_DIR.glob("*.xlsx"))
    if not candidates:
        sys.exit(f"no .xlsx in {DRIVEN_DIR}")
    if len(candidates) > 1:
        print(f"multiple xlsx found, using most recent: {candidates[-1].name}", file=sys.stderr)
    return candidates[-1]


def fmt(value, field: str):
    if value is None:
        return ""
    if field in DATE_COLS:
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        # openpyxl should parse date-styled cells to datetime; if it comes
        # through as a raw serial, fall back.
        if isinstance(value, (int, float)):
            base = datetime(1899, 12, 30)
            return (base.fromordinal(base.toordinal() + int(value))).strftime("%Y-%m-%d")
    return value


def main() -> None:
    src = pick_input(sys.argv)
    print(f"reading  {src}", file=sys.stderr)
    wb = load_workbook(src, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        sys.exit("empty sheet")

    resolved = []
    missing = []
    for h in raw_headers:
        key = HEADER_MAP.get(str(h).strip()) if h is not None else None
        if key is None:
            missing.append(h)
        resolved.append(key)
    if missing:
        sys.exit(f"unmapped headers: {missing}")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with OUT.open("w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow([HEADER_MAP[str(h).strip()] for h in raw_headers])
        for row in rows_iter:
            if all(v is None for v in row):
                continue
            w.writerow([fmt(v, resolved[i]) for i, v in enumerate(row)])
            count += 1

    print(f"wrote    {OUT}  ({count:,} rows)", file=sys.stderr)


if __name__ == "__main__":
    main()
